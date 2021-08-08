import os
import openpyxl
import search_file
import pathlib


def execute(target_dir, exclusion_dir, sheet_name, work_file, header_record):
    list = search_file.execute(target_dir, exclusion_dir.split(','))
    new_work_book = openpyxl.Workbook()
    new_work_sheet = new_work_book["Sheet"]
    new_work_sheet.title = sheet_name
    record_count = 1
    is_configured_header = False
    for excel_file_path in list:
        work_book = openpyxl.load_workbook(excel_file_path)
        work_sheet = work_book[sheet_name]
        target_record_list = [
            cell.row for cell in work_sheet["A:A"] if cell.value is not None]
        a_max_row = max(target_record_list, key=lambda record: record)
        if not is_configured_header:

            for header in work_sheet.iter_rows(min_row=header_record, max_row=header_record):
                new_work_sheet.cell(record_count, 1).value = 'No'
                new_work_sheet.cell(record_count, 2).value = 'FilePath'

                copying_record(new_work_sheet, record_count, header)
                record_count += 1
            is_configured_header = True

        path_file = pathlib.Path(excel_file_path)

        for record in work_sheet.iter_rows(min_row=header_record+1, max_row=a_max_row,  max_col=work_sheet.max_column):
            new_work_sheet.cell(record_count, 1).value = record_count-1
            new_work_sheet.cell(record_count, 2).value = excel_file_path
            new_work_sheet.cell(record_count, 2).hyperlink = excel_file_path
            copying_record(new_work_sheet, record_count, record)
            record_count += 1
    new_work_book.save(work_file)


def copying_record(new_work_sheet, record_count, record):
    for cell in record:
        new_work_sheet.cell(record_count, cell.column +
                            2).alignment = openpyxl.styles.Alignment(wrapText=True)
        new_work_sheet.cell(record_count, cell.column+2).value = cell.value


target_dir = "C:/Users/tnaka/OneDrive/デスクトップ/hoge/"
exclusion_dir = '対象外,たいしょうがい'
sheet_name = 'hoge'
work_file = 'C:/Users/tnaka/OneDrive/デスクトップ/temp.xlsx'
header_record = 2
execute(target_dir, exclusion_dir, sheet_name, work_file, header_record)
