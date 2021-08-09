
import openpyxl
from itertools import groupby


def execute(sheet_name, work_file, header_record):
    index_work_book = openpyxl.load_workbook(work_file)
    index_work_sheet = index_work_book[sheet_name]
    for key, record_list in groupby(index_work_sheet.iter_rows(min_row=2), key=lambda record: record[1].value):
        target_work_book = openpyxl.load_workbook(key)
        target_work_sheet = target_work_book[sheet_name]
        target_record_list = [
            cell.row for cell in target_work_sheet["A:A"] if cell.value is not None]
        a_max_row = max(target_record_list, key=lambda record: record)
        target_list = list(record_list)
        if len(target_list)+header_record > a_max_row:
            insert_count = len(target_list)+header_record - a_max_row
            for i in range(0, insert_count):
                target_work_sheet.insert_rows(a_max_row)
        if len(target_list)+header_record < a_max_row:
            delete_count = len(target_list)+header_record - a_max_row
            target_work_sheet.delete_rows(
                idx=header_record+1, amount=delete_count)
        for index, record in enumerate(target_list):
            target_work_sheet.row_dimensions[index+header_record +
                                             1].height = index_work_sheet.row_dimensions[record[0].row].height
            for cell in record[2:]:
                target_work_sheet.cell(index+header_record+1,
                                       cell.column - 2).value = cell.value
                target_work_sheet.cell(index+header_record+1,
                                       cell.column - 2).font = cell.font._StyleProxy__target
                target_work_sheet.cell(index+header_record+1,
                                       cell.column - 2).alignment = cell.alignment._StyleProxy__target
        target_work_book.save(key)
